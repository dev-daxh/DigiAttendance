from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import mysql.connector
import time
import openpyxl
from tkinter.simpledialog import askinteger
import os

main = Tk()
main.geometry("925x500+300+200")
main.config(bg="#fff")
main.resizable(False, False)
main.title("Main Page")

# Initialize MySQL connection
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="root@123",
    database="qr"  # Change database name accordingly
)
mycur = mydb.cursor()

# Paths to Excel files for each course
course_files = {
    "cm201g": r"D:\Python Project\Digiattendance\Excel Data\cm201g.xlsx",
    "cm202g": r"D:\Python Project\Digiattendance\Excel Data\cm202g.xlsx",
    "cm203g": r"D:\Python Project\Digiattendance\Excel Data\cm203g.xlsx",
    "cm204g": r"D:\Python Project\Digiattendance\Excel Data\cm204g.xlsx",
    "cm205g": r"D:\Python Project\Digiattendance\Excel Data\cm205g.xlsx",
}

# Function to mark attendance in Excel
def mark_attendance_in_excel(course_code, enrollment_number):
    file_path = course_files.get(course_code.lower())
    
    if not file_path:
        messagebox.showerror("Error", f"Invalid course code {course_code}")
        return

    if not os.path.exists(file_path):
        messagebox.showerror("Error", f"Excel file for course {course_code} not found at {file_path}!")
        return

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except PermissionError:
        messagebox.showerror("Error", f"Permission denied: Unable to open {file_path}")
        return

    # Find the first empty row
    empty_row = 2
    while sheet.cell(row=empty_row, column=1).value is not None:
        empty_row += 1

    # Store enrollment number and current date and time in the empty row
    sheet.cell(row=empty_row, column=1, value=enrollment_number)
    sheet.cell(row=empty_row, column=2, value="P")
    sheet.cell(row=empty_row, column=3, value=time.strftime("%Y-%m-%d"))
    sheet.cell(row=empty_row, column=4, value=time.strftime("%H:%M:%S"))

    try:
        workbook.save(file_path)
        print("Attendance marked successfully in Excel.")
    except PermissionError:
        messagebox.showerror("Error", f"Permission denied: Unable to save {file_path}")
        return

# Function to fetch and update attendance
def excel_ope():
    global enrollment, course
    enrollment = askinteger("confirmation", "Enter your Enrollment Number:")
    if not enrollment:
        return  # User canceled the input

    table_name = course.lower()
    query = f"SELECT * FROM {table_name} WHERE qr_code = %s"
    mycur.execute(query, (qr_code_data,))
    row = mycur.fetchone()

    if row is None:
        messagebox.showerror("Error", "Invalid QR code for the selected course! Scan properly")
    else:
        mark_attendance_in_excel(table_name, enrollment)
        mark_attendance_in_database(enrollment, table_name)

# Function to mark attendance in the database
def mark_attendance_in_database(enroll, course):
    try:
        mycur.execute("INSERT INTO attendance (enroll, course) VALUES (%s, %s)", (enroll, course))
        mydb.commit()
        messagebox.showinfo("Congratulations", "Attendance Done Successfully!!")
    except mysql.connector.Error as err:
        messagebox.showerror("Error", f"Failed to mark attendance in database: {err}")

def qr_validation():
    global qr_code_data, course
    qr_code_data = user.get()
    if qr_code_data == "Scan the QR and Enter the code":
        messagebox.showerror("error", "Input some unique text!")
        return

    course = final_course.get()
    if course == "Select a Course":
        messagebox.showerror("error", "Select a Course!")
        return
    else:
        excel_ope()

def curtime():
    timeVar = time.strftime("%I:%M:%S %p\n %A \n %x")
    clock.config(text=timeVar)
    clock.after(200, curtime)

img = PhotoImage(file="images/welcome_main.png")
Label(main, image=img, bg="white").place(x=1, y=5)
clock = Label(main, font=("Calibri", 13), bg="#57a1f8", fg="white")
clock.place(x=820, y=5)

# Course dropdown menu
course_codes = ["CM201G", "CM202G", "CM203G", "CM204G", "CM205G"]
final_course = StringVar()
course_dropdown = ttk.Combobox(main, textvariable=final_course)
course_dropdown["values"] = course_codes
course_dropdown.place(x=620, y=150)
course_dropdown.set("Select a Course")

def on_enter(e):
    user.delete(0, "end")

def on_leave(e):
    name = user.get()
    if name == "":
        user.insert(0, "Scan the QR and Enter the code")

user = Entry(main, width=35, fg="black", border=0, bg="white", font=("Microsoft YaHei UILight", 10))
user.place(x=620, y=210)
user.insert(0, "Scan the QR and Enter the code")
user.bind("<FocusIn>", on_enter)
user.bind("<FocusOut>", on_leave)

Frame(main, width=200, height=2, bg="black").place(x=620, y=235)
Button(main, width=39, pady=7, text="Submit", bg="#57a1f8", fg="white", border=0, command=qr_validation).place(x=610, y=280)
curtime()
Button(main, width=9, pady=7, text="Log Out", bg="#57a1f8", fg="white", border=0).place(x=800, y=450)

main.mainloop()
